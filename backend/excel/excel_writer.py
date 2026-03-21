"""Generate Excel output from extracted freight data — no template required."""

from __future__ import annotations

import logging
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from mapping.field_mapper import map_rate_rows, map_origin_arb_rows, map_dest_arb_rows

logger = logging.getLogger("excel_writer")

# Excel serial date epoch
_EXCEL_EPOCH = date(1899, 12, 30)

# ── Column definitions (field_name, display_header) ──────────────────────────

RATES_COLUMNS = [
    ("carrier",              "Carrier"),
    ("contract_id",          "Contract ID"),
    ("effective_date",       "Effective Date"),
    ("expiration_date",      "Expiration Date"),
    ("commodity",            "Commodity"),
    ("origin_city",          "Origin City"),
    ("origin_via_city",      "Origin Via City"),
    ("destination_city",     "Destination City"),
    ("destination_via_city", "Destination Via City"),
    ("service",              "Service"),
    ("remarks",              "Remarks"),
    ("scope",                "Scope"),
    ("base_rate_20",         "BaseRate 20"),
    ("base_rate_40",         "BaseRate 40"),
    ("base_rate_40h",        "BaseRate 40H"),
    ("base_rate_45",         "BaseRate 45"),
    ("ams_china_japan",      "AMS"),
    ("hea_heavy_surcharge",  "HEA"),
    ("agw",                  "AGW"),
    ("rds_red_sea",          "RDS"),
    ("meo",                  "MEO"),
    ("pef",                  "PEF"),
    ("reefer_rate_20",       "Reefer 20"),
    ("reefer_rate_40",       "Reefer 40"),
    ("reefer_rate_40h",      "Reefer 40H"),
    ("reefer_rate_nor40",    "Non Operating Reefer"),
]

ORIGIN_ARB_COLUMNS = [
    ("carrier",              "Carrier"),
    ("contract_id",          "Contract ID"),
    ("effective_date",       "Effective Date"),
    ("expiration_date",      "Expiration Date"),
    ("commodity",            "Commodity"),
    ("origin_city",          "Origin City"),
    ("origin_via_city",      "Origin Via City"),
    ("service",              "Service"),
    ("remarks",              "Remarks"),
    ("scope",                "Scope"),
    ("base_rate_20",         "BaseRate 20"),
    ("base_rate_40",         "BaseRate 40"),
    ("base_rate_40h",        "BaseRate 40H"),
    ("base_rate_45",         "BaseRate 45"),
    ("agw_20",               "AGW 20"),
    ("agw_40",               "AGW 40"),
    ("agw_45",               "AGW 45"),
]

DEST_ARB_COLUMNS = [
    ("carrier",              "Carrier"),
    ("contract_id",          "Contract ID"),
    ("effective_date",       "Effective Date"),
    ("expiration_date",      "Expiration Date"),
    ("commodity",            "Commodity"),
    ("destination_city",     "Destination City"),
    ("destination_via_city", "Destination Via City"),
    ("service",              "Service"),
    ("remarks",              "Remarks"),
    ("scope",                "Scope"),
    ("base_rate_20",         "BaseRate 20"),
    ("base_rate_40",         "BaseRate 40"),
    ("base_rate_40h",        "BaseRate 40H"),
    ("base_rate_45",         "BaseRate 45"),
]

# ── Styling constants ────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_FONT = Font(name="Calibri", size=10)
_DATA_ALIGN = Alignment(vertical="top")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


def _to_excel_date(value: str) -> Optional[int]:
    """Convert date string to Excel serial number."""
    if not value:
        return None
    formats = ["%d %b %Y", "%d %B %Y", "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"]
    for fmt in formats:
        try:
            d = datetime.strptime(value.strip(), fmt).date()
            return (d - _EXCEL_EPOCH).days
        except ValueError:
            continue
    return None


def _cell_value(obj, field: str):
    val = getattr(obj, field, None)
    if field in ("effective_date", "expiration_date"):
        return _to_excel_date(val) if isinstance(val, str) else val
    return val


def _create_sheet(wb: Workbook, title: str, columns: list[tuple[str, str]],
                  rows: list) -> int:
    """Create a styled sheet with headers and data. Returns row count written."""
    ws = wb.create_sheet(title=title)

    # Write header row
    for col_idx, (field, header) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # Write data rows
    for i, row_obj in enumerate(rows):
        excel_row = 2 + i
        for col_idx, (field, _header) in enumerate(columns, start=1):
            val = _cell_value(row_obj, field)
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = _DATA_ALIGN
            cell.border = _THIN_BORDER
            # Alternate row shading
            if i % 2 == 1:
                cell.fill = _ALT_ROW_FILL

    # Auto-fit column widths (approximate)
    for col_idx, (field, header) in enumerate(columns, start=1):
        # Sample first 50 data rows for width estimation
        max_len = len(header)
        for row_idx in range(2, min(2 + len(rows), 52)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        # Clamp between 8 and 30 characters
        width = min(max(max_len + 2, 8), 30)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if rows:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{1 + len(rows)}"

    return len(rows)


def write_excel(structured: dict, output_path: str) -> None:
    """
    Generate a fresh Excel workbook with extracted freight data.

    Args:
        structured: output from the NLP/AI extraction pipeline
        output_path: where to save the .xlsx result
    """
    # Map to typed dataclasses
    rate_rows = map_rate_rows(structured)
    origin_arb_rows = map_origin_arb_rows(structured)
    dest_arb_rows = map_dest_arb_rows(structured)

    wb = Workbook()
    # Remove default sheet created by openpyxl
    wb.remove(wb.active)

    # ── Rates sheet ──────────────────────────────────────────────────────────
    rates_written = _create_sheet(wb, "Rates", RATES_COLUMNS, rate_rows)

    # ── Origin Arbitraries sheet ─────────────────────────────────────────────
    origin_written = 0
    if origin_arb_rows:
        origin_written = _create_sheet(wb, "Origin Arbitraries",
                                       ORIGIN_ARB_COLUMNS, origin_arb_rows)

    # ── Destination Arbitraries sheet ────────────────────────────────────────
    dest_written = 0
    if dest_arb_rows:
        dest_written = _create_sheet(wb, "Destination Arbitraries",
                                     DEST_ARB_COLUMNS, dest_arb_rows)

    wb.save(output_path)
    logger.info(f"[excel_writer] Saved {rates_written} rate rows, "
                f"{origin_written} origin arbs, "
                f"{dest_written} dest arbs -> {output_path}")
